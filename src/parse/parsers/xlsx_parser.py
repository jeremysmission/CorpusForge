"""
XLSX parser -- reads modern Excel (.xlsx) spreadsheets.

Plain English: opens an Excel workbook and turns each sheet into text
the Forge pipeline can chunk. Each sheet gets a ``[SHEET] <name>``
header so operators can trace chunks back to the right tab. The first
non-empty row is treated as the column header, and the remaining rows
are rendered as ``Header: value, Header: value`` pairs so an engineer
reading a retrieved chunk can still tell which column a value came from.

Uses openpyxl in read-only mode for speed. Hard cap of 100,000 rows
per sheet protects against runaway spreadsheets.

Ported from V1 (src/parsers/office_xlsx_parser.py).
"""

from __future__ import annotations

import logging
from pathlib import Path

from src.parse.parsers.txt_parser import ParsedDocument

logger = logging.getLogger(__name__)

_MAX_ROWS = 100_000  # safety cap: stop reading a sheet after this many rows


class XlsxParser:
    """Extract text from modern Excel .xlsx workbooks, sheet by sheet."""

    def parse(self, file_path: Path) -> ParsedDocument:
        """Open an .xlsx file and return all sheets rendered as key-value text."""
        path = Path(file_path)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            sheets = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_lines = [f"[SHEET] {sheet_name}"]
                headers = None
                row_count = 0

                for row in ws.iter_rows(values_only=True):
                    row_count += 1
                    if row_count > _MAX_ROWS:
                        break

                    cells = [self._cell_to_str(c) for c in row]
                    if all(not c for c in cells):
                        continue

                    if headers is None:
                        headers = cells
                        sheet_lines.append(" | ".join(headers))
                        continue

                    # Key-value representation
                    pairs = []
                    for h, v in zip(headers, cells):
                        if v:
                            pairs.append(f"{h}: {v}")
                    if pairs:
                        sheet_lines.append(", ".join(pairs))

                sheets.append("\n".join(sheet_lines))
            wb.close()
            text = "\n\n".join(sheets)
        except Exception as e:
            logger.error("XLSX parse failed for %s: %s", path.name, e)
            text = ""

        return ParsedDocument(
            source_path=str(path),
            text=text,
            parse_quality=0.9 if text.strip() else 0.0,
            file_ext=".xlsx",
            file_size=path.stat().st_size,
        )

    def _cell_to_str(self, value) -> str:
        if value is None:
            return ""
        if hasattr(value, "isoformat"):
            return value.isoformat()
        return str(value).strip()
